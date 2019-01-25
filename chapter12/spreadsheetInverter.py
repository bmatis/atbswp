#!/usr/bin/env python3
# spreadsheetInverter.py - inverts the rows and columns of a spreadsheet
# example, value at row 5 column 3 will end up in row 3 column 5

import openpyxl, sys

# get the desired spreadsheet from the command line argument
sourceFile = sys.argv[1]

# get source workbook
sourceWb = openpyxl.load_workbook(sourceFile)
sourceSheet = sourceWb.active

# create destination workbook
newWb = openpyxl.Workbook()
newSheet = newWb.active

for rowNum in range(1, sourceSheet.max_row + 1):
    for colNum in range(1, sourceSheet.max_column + 1):
        sourceCell = sourceSheet.cell(row=rowNum, column=colNum)
        destCell = newSheet.cell(row=colNum, column=rowNum)
        destCell.value = sourceCell.value

# save the new workbook
baseFile = sourceFile.rsplit('.', 1)[0] # strips out the file extension
newWb.save(baseFile + '_inverted.xlsx')