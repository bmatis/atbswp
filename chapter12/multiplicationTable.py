#!/usr/bin/env python3
# multiplicationTable.py - takes a number (N) from command line and creates
# an N x N mulitplication table in an Excel spreadsheet
# example: multiplicationTable.py 6 to create a 6 x 6 multiplication table

import sys, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# get command line argument
try:
    value = int(sys.argv[1])
except:
    print('Error: invalid command line argument. Please provide an int.')
    exit()

# open an Excel workbook and get active sheet
wb = openpyxl.Workbook()
sheet = wb.active

# create a font object for cells that should be bold
boldFont = Font(bold=True)

# generate x axis legend
for col in range(1, value+1):
    col_letter = get_column_letter(col+1)
    cell = col_letter + str(1)
    sheet[cell] = col
    sheet[cell].font = boldFont

# generate y axis legend
for row in range(1, value+1):
    cell = 'A' + str(row + 1)
    sheet[cell] = row
    sheet[cell].font = boldFont

# generate multiplication table
for row in range(1, value+1):
    for col in range(1, value+1):
        col_letter = get_column_letter(col+1)
        sheet[col_letter + str(row+1)] = int(row) * int(col)

# save the workbook
wb.save('multiplicationTable_' + str(value) + '.xlsx')