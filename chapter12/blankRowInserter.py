#!/usr/bin/env python3
# blankRowInserter.py - starting at row N, will insert M number of blank rows
# into an excel spreadsheet
# example usage: blankRowInserter.py 3 2 mySpreadsheet.xlsx

import openpyxl, sys

# get the command line arguments
splitRow = int(sys.argv[1])
blankRows = int(sys.argv[2])
sourceFile = sys.argv[3]

# get source workbook
sourceWb = openpyxl.load_workbook(sourceFile)
sourceSheet = sourceWb.active

# create destination workbook
newWb = openpyxl.Workbook()
newSheet = newWb.active

def copyRow(sourceRowNum, destRowNum):
    """Copy a source row to a destination row in the new spreadsheet."""
    for colNum in range(1, sourceSheet.max_column + 1):
        destCell = newSheet.cell(row=destRowNum, column=colNum)
        sourceCell = sourceSheet.cell(row=sourceRowNum, column=colNum)
        destCell.value = sourceCell.value

# copy the first rows up until the split
for rowNum in range(1, splitRow):
    copyRow(rowNum, rowNum)

# copy the rest of the workbook, but with blank rows at the split
for rowNum in range(splitRow, sourceSheet.max_row + 1):
    copyRow(rowNum, rowNum + blankRows)

# save the new workbook
baseFile = sourceFile.rsplit('.', 1)[0] # strips out the file extension
newWb.save(baseFile + '_withBlankRows.xlsx')